import csv
import openpyxl
ncsv=input("Введите название файла - ")
with open(ncsv,encoding='utf-8') as file:
    rl=list(csv.reader(file))
    fs=rl[0]
    sr=sorted(sorted(sorted(rl[1:], key=lambda x:x[2]), key=lambda x:x[1]),key=lambda x:x[3])
wb = openpyxl.Workbook()
nxls=input("Введите название файла - ")
wb.save(nxls)
ws =wb.active
ws.append(tuple(fs))
for i in sr:
    ws.append(tuple(i))
su = 0
for i in range(2,ws.max_row+1):
    su+=int(ws.cell(i,5).value)
j = ws.max_row+1
ws.cell(j,1).value='Итого'
ws.cell(j,5).value=su
wb.save(nxls)