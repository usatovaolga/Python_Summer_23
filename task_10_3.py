import openpyxl
wb = openpyxl.load_workbook('task_10_2.xlsx')
ws = wb['total']
sp=[]
def median(s):
    if len(s)%2==0:
        m = (s[len(s)//2]+s[len(s)//2-1])/2
    else:
        m =s[len(s)//2]
    return m
for i in range(1,ws.max_row):
    sp.append(ws.cell(i, 2).value)
print(sp)
d={'Минимальное значение':min(sp),
   'Максимальное значение':max(sp),
   'Среднее арифметическое':sum(sp)/len(sp),
   'Медиана':median(sorted(sp))}
print(d)
wb.create_sheet('total2')
wss = wb['total2']
n=1
for k,v in dict(sorted(d.items(), key=lambda x: x[0])).items():
    wss.cell(n,1).value = k
    wss.cell(n,2).value = v
    n+=1
wb.save('task_10_2.xlsx')