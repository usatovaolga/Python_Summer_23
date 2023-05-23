import openpyxl
wb = openpyxl.load_workbook('task_10_2.xlsx')
s = wb.sheetnames
d={}
for y in s:
    ws = wb[y]
    for i in range(1,ws.max_row+1):#до последней строки где есть какое-то значение
        d[ws.cell(i,1).value]=d.get(ws.cell(i, 1).value,0)+ws.cell(i, 2).value
print(d)
wb.create_sheet('total')
wss = wb['total']
n=1
for k,v in dict(sorted(d.items(), key=lambda x: x[0])).items():
    wss.cell(n,1).value = k
    wss.cell(n,2).value = v
    n+=1
su = 0
for i in range(1,wss.max_row+1):
    su+=wss.cell(i,2).value
j = wss.max_row+1
wss.cell(j,1).value='Итого'
wss.cell(j,2).value=su
wb.save('task_10_2.xlsx')